from time import sleep
import openpyxl.workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.remote.webelement import WebElement
import openpyxl, json, re


def wait_url(driver: webdriver.Chrome, url: str):
    while True:
        cur_url = driver.current_url
        if cur_url == url:
            break
        sleep(0.1)

def find_element(driver: webdriver.Chrome, whichBy, unique: str) -> WebElement:
    while True:
        try:
            element = driver.find_element(whichBy, unique)
            break
        except:
            pass
        sleep(1)
    return element

def find_elements(driver : webdriver.Chrome, whichBy, unique: str) -> list[WebElement]:
    while True:
        try:
            elements = driver.find_elements(whichBy, unique)
            break
        except:
            pass
        sleep(1)
    return elements

wb = openpyxl.Workbook()

ws = wb.active

ws.merge_cells('F1:K1')
ws.merge_cells('L1:M1')
ws.merge_cells('N1:S1')
ws.merge_cells('U1:V1')
ws.merge_cells('A1:A2')
ws.merge_cells('B1:B2')
ws.merge_cells('C1:C2')

ws['F1'] = 'MEDIDAS'
ws['L1'] = 'COLORS'
ws['N1'] = 'MATERIALES Y CARACTERÍSTICAS'
ws['T1'] = 'OTROS BENEFICIOS'
ws['U1'] = 'GARANTÍA Y COMPONENTES'

ws['A1'] = "Categoría"
ws['B1'] = "Línea"
ws['C1'] = "Sublinea"
ws['D2'] = "Atributos1"
ws['E2'] = "Atributos2"
ws['F2'] = "Atributos3"
ws['G2'] = "Atributos4"
ws['H2'] = "Atributos5"
ws['I2'] = "Atributos6"
ws['J2'] = "Atributos7"
ws['K2'] = "Atributos8"
ws['L2'] = "Atributos9"
ws['M2'] = "Atributos10"
ws['N2'] = "Atributos11"
ws['O2'] = "Atributos11"
ws['P2'] = "Atributos11"
ws['Q2'] = "Atributos11"
ws['R2'] = "Atributos11"
ws['S2'] = "Atributos11"
ws['T2'] = "Atributos11"
ws['U2'] = "Atributos11"
ws['V2'] = "Atributos11"
ws['W2'] = "Atributos11"

wb.save(f'jamar.xlsx')

urls = [
  "https://www.jamar.com/collections/sofas",
  "https://www.jamar.com/collections/sofa-en-l",
  "https://www.jamar.com/collections/juego-de-sala",
  "https://www.jamar.com/collections/sofa-camas",
  "https://www.jamar.com/collections/sillon-reclinable",
  "https://www.jamar.com/collections/poltronas-y-puffs",   
  "https://www.jamar.com/collections/mesas-de-centro-y-auxiliares",   
  "https://www.jamar.com/collections/mesas-de-marmol",   
  "https://www.jamar.com/collections/types?query=muebles-para-tv&category=WL2-034",   
  "https://www.jamar.com/collections/espejo",   
  "https://www.jamar.com/collections/consolas",   
  "https://www.jamar.com/collections/mecedoras",   
  "https://www.jamar.com/collections/juegos-de-comedor",   
  "https://www.jamar.com/collections/sillas-de-comedor",   
  "https://www.jamar.com/collections/mesa-comedor",   
  "https://www.jamar.com/collections/bifes-y-vitrinas",   
  "https://www.jamar.com/collections/sillas-para-barra",   
  "https://www.jamar.com/collections/camas",   
  "https://www.jamar.com/collections/juegos-de-alcobas",   
  "https://www.jamar.com/collections/tocadores",   
  "https://jamar.com/collections/mesas-de-noche",   
  "https://www.jamar.com/collections/comodas",   
  "https://www.jamar.com/collections/cabeceros",   
  "https://www.jamar.com/collections/bases-de-cama",   
  "https://www.jamar.com/collections/puffs",   
  "https://www.jamar.com/collections/colchones-king",   
  "https://www.jamar.com/collections/colchones-queen",   
  "https://www.jamar.com/collections/colchones-dobles",   
  "https://www.jamar.com/collections/colchones-semi-dobles",   
  "https://www.jamar.com/collections/colchones-sencillos",   
  "https://www.jamar.com/collections/escritorio",   
  "https://www.jamar.com/collections/biblioteca",   
  "https://www.jamar.com/collections/sillas-de-escritorio",   
  "https://www.jamar.com/collections/muebles-infantiles",  
]
match_num = 0

driver = webdriver.Chrome()
driver.maximize_window()

for url in urls:
    driver.get(url)
    wait_url(driver, url)

    try:
        driver.find_element(By.CLASS_NAME, "mfp-close").click()
    except:
        pass

    driver.execute_script("window.scrollTo(0, document.body.scrollHeight/2)")
    sleep(0.3)
    driver.execute_script("window.scrollTo(0, document.body.scrollHeight*3/4)")
    sleep(5)

    while True:
        try:
            driver.find_element(By.CLASS_NAME, "js-view-more-products").click()
            sleep(5)
            try:
                driver.find_element(By.CLASS_NAME, "fancyboxPop-close-pop").click()
            except:
                pass
        except:
            break
        sleep(0.1)

    products = find_elements(driver, By.CLASS_NAME, "card-views")
    product_urls = []
    
    for product in products:
        product_url = product.find_element(By.TAG_NAME, "a").get_attribute("href")
        product_urls.append(product_url)
        print(product_url)

    for product_url in product_urls:
        match_num += 1
        workbook = openpyxl.load_workbook("jamar.xlsx")
        sheet = workbook['Sheet']
        driver.get(product_url)
        sleep(0.5)
        try:
            driver.find_element(By.CLASS_NAME, "contenedor-parrafo-boton-registra-404")
            pass
        except:
            product_num = re.findall(r'\d+', find_element(driver, By.CLASS_NAME, "txt-sku").text)[0]
            print(product_num)
            sheet[f'D{match_num +2}'] = product_num
            product_name = find_element(driver, By.CLASS_NAME, "product-title").text.split("\n")[0]
            # product_subname = find_element(driver, By.CLASS_NAME, "product-title").text.split("\n")[1]
            print(product_name)
            sheet[f'E{match_num +2}'] = product_name
            # detail dropdown click!
            find_element(driver, By.CLASS_NAME, "container-detail-product").find_elements(By.TAG_NAME, "details")[1].click()
            print("Clicked!")
            sleep(1)
            try:
                driver.find_element(By.CLASS_NAME, "fancyboxPop-close-pop").click()
            except:
                pass

            # Get data of product
            product_detail = find_element(driver, By.CLASS_NAME, "container-visor-detail").find_element(By.CLASS_NAME, "container-detail-product").find_elements(By.TAG_NAME, "details")[1].find_element(By.CLASS_NAME, "product-description").find_elements(By.XPATH, ".//div/table")
            print(len(product_detail))
            # Product measurements
            try:
                product_measurements = product_detail[0].find_element(By.TAG_NAME,"tbody").find_elements(By.XPATH, ".//tr")    
                for product_measurement in product_measurements:
                    if product_measurement.find_elements(By.XPATH, ".//td")[0].text == "Alto":
                        product_high = product_measurement.find_elements(By.XPATH, ".//td")[1].text
                        sheet[f'F{match_num +2}'] = product_high
                    elif product_measurement.find_elements(By.XPATH, ".//td")[0].text == "Ancho":
                        product_board = product_measurement.find_elements(By.XPATH, ".//td")[1].text
                        sheet[f'G{match_num +2}'] = product_board
                    elif product_measurement.find_elements(By.XPATH, ".//td")[0].text == "Profundo":
                        product_deep = product_measurement.find_elements(By.XPATH, ".//td")[1].text
                        sheet[f'H{match_num +2}'] = product_deep
                    elif product_measurement.find_elements(By.XPATH, ".//td")[0].text == "Medidas Sofá 1 (Alto X Ancho X Profundidad)":
                        product_sofa_size = product_measurement.find_elements(By.XPATH, ".//td")[1].text
                        sheet[f'I{match_num +2}'] = product_sofa_size
                    elif product_measurement.find_elements(By.XPATH, ".//td")[0].text == "Medidas Butaco (A) (Alto X Ancho X Profundidad)":
                        product_seat_size = product_measurement.find_elements(By.XPATH, ".//td")[1].text
                        sheet[f'J{match_num +2}'] = product_seat_size
                    elif product_measurement.find_elements(By.XPATH, ".//td")[0].text == "Medidas Puff (Alto X Ancho X Profundidad)":
                        product_puff_size = product_measurement.find_elements(By.XPATH, ".//td")[1].text
                        sheet[f'K{match_num +2}'] = product_puff_size
            except:
                pass

            # Product Colors
            try:
                product_colors = product_detail[1].find_element(By.TAG_NAME,"tbody").find_elements(By.XPATH, ".//tr")
                for product_color in product_colors:
                    if product_color.find_elements(By.XPATH, ".//td")[0].text == "Color Estructura":
                        product_structure_color = product_color.find_elements(By.XPATH, ".//td")[1].text
                        sheet[f'L{match_num +2}'] = product_structure_color
                    elif product_color.find_elements(By.XPATH, ".//td")[0].text == "Color Tapiz":
                        product_color_tapestry1 = product_color.find_elements(By.XPATH, ".//td")[1].text
                        sheet[f'M{match_num +2}'] = product_color_tapestry1
            except:
                pass

            # Product Materials
            try:
                product_materials = product_detail[2].find_element(By.TAG_NAME,"tbody").find_elements(By.XPATH, ".//tr")
                for product_material in product_materials:
                    if product_material.find_elements(By.XPATH, ".//td")[0].text == "Material De La Estructura":
                        product_structure_material = product_material.find_elements(By.XPATH, ".//td")[1].text
                        sheet[f'N{match_num +2}'] = product_structure_material
                    elif product_material.find_elements(By.XPATH, ".//td")[0].text == "Material Del Relleno":
                        product_filling_material = product_material.find_elements(By.XPATH, ".//td")[1].text
                        sheet[f'O{match_num +2}'] = product_filling_material
                    elif product_material.find_elements(By.XPATH, ".//td")[0].text == "Composición":
                        product_composition_material = product_material.find_elements(By.XPATH, ".//td")[1].text
                        sheet[f'P{match_num +2}'] = product_composition_material
                    elif product_material.find_elements(By.XPATH, ".//td")[0].text == "Tipo De Pata":
                        product_leg_material = product_material.find_elements(By.XPATH, ".//td")[1].text
                        sheet[f'Q{match_num +2}'] = product_leg_material
                    elif product_material.find_elements(By.XPATH, ".//td")[0].text == "Tipo de Tela":
                        product_fabric_material = product_material.find_elements(By.XPATH, ".//td")[1].text
                        sheet[f'R{match_num +2}'] = product_fabric_material
                    elif product_material.find_elements(By.XPATH, ".//td")[0].text == "Patas Desmontables":
                        product_detachable_material = product_material.find_elements(By.XPATH, ".//td")[1].text
                        sheet[f'S{match_num +2}'] = product_detachable_material
            except:
                pass

            # Produt Benefits
            try:
                product_benefits = product_detail[3].find_element(By.TAG_NAME,"tbody").find_elements(By.XPATH, ".//tr")
                for product_benefit in product_benefits:
                    if product_benefit.find_elements(By.XPATH, ".//td")[0].text == "Uso":
                        product_benefit_use = product_benefit.find_elements(By.XPATH, ".//td")[1].text
                        sheet[f'T{match_num +2}'] = product_benefit_use
            except:
                pass

            # Product Warranty
            try:
                product_warranties = product_detail[4].find_element(By.TAG_NAME,"tbody").find_elements(By.XPATH, ".//tr")
                print(len(product_warranties))
                for product_warranty in product_warranties:
                    if product_warranty.find_elements(By.XPATH, ".//td")[0].text == "Garantía":
                        product_warranty_warranty = product_warranty.find_elements(By.XPATH, ".//td")[1].text
                        sheet[f'U{match_num +2}'] = product_warranty_warranty
                    elif product_warranty.find_elements(By.XPATH, ".//td")[0].text == "Incluye":
                        product_warranty_includes = product_warranty.find_elements(By.XPATH, ".//td")[1].text
                        sheet[f'V{match_num +2}'] = product_warranty_includes
            except:
                pass
            image_url = find_element(driver, By.CLASS_NAME, "container-visor-product").find_element(By.TAG_NAME, "img").get_attribute("src")
            sheet[f'W{match_num + 1}'] = image_url
            workbook.save('jamar.xlsx')
        
    match_num +=2